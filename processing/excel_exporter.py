"""
Workers Compensation Data Exporter - Excel Output
Imports and properly formats combined workers comp CSV data into the Excel spreadsheet template.
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
from datetime import datetime, timedelta
import re


def parse_employee_name(full_name):
    """
    Parse employee name into first and last name.
    Expected format: "Last , First Middle" or "Last, First"
    """
    if pd.isna(full_name) or full_name == "":
        return "", ""

    full_name = str(full_name).strip()
    full_name = re.sub(r'\s+', ' ', full_name)

    if ',' in full_name:
        parts = full_name.split(',', 1)
        last_name = parts[0].strip()
        first_name = parts[1].strip() if len(parts) > 1 else ""
        return first_name, last_name
    else:
        return "", full_name


def get_class_code_mapping():
    """
    Define the high-wage to low-wage class code mappings for California Workers' Comp.
    """
    return {
        5432: 5403,
        543221: 540321,
    }


def get_wage_thresholds():
    """
    Define wage thresholds for dual-wage classifications.
    """
    return {
        'carpentry': {
            'high_code': 5432,
            'low_code': 5403,
            'threshold': 39.00,
            'name': 'Carpentry'
        },
        'painting': {
            'high_code': 5482,
            'low_code': None,
            'threshold': 31.00,
            'name': 'Painting'
        },
        'plastering': {
            'high_code': 5485,
            'low_code': None,
            'threshold': 36.00,
            'name': 'Plastering/Stucco'
        },
        'roofing': {
            'high_code': 5553,
            'low_code': None,
            'threshold': 27.00,
            'name': 'Roofing'
        }
    }


def validate_class_code_by_wage(df):
    """
    Validate that each transaction is properly classified based on actual hourly rate.
    """
    thresholds = get_wage_thresholds()

    high_codes = {}
    low_codes = {}

    for trade_key, trade_info in thresholds.items():
        high_code = trade_info['high_code']
        low_code = trade_info['low_code']
        threshold = trade_info['threshold']
        name = trade_info['name']

        if high_code:
            high_codes[high_code] = {'threshold': threshold, 'name': name, 'low_code': low_code}
        if low_code:
            low_codes[low_code] = {'threshold': threshold, 'name': name, 'high_code': high_code}

    regular_wage_types = ['REG', 'VAC', 'SICK', 'DBA', 'SUPP', 'SAL', 'OSAL', 'PWREG']
    validation_mask = df['Earn Type'].str.upper().str.strip().isin(regular_wage_types)

    misclassifications = []

    for idx in df[validation_mask].index:
        try:
            earnings = float(df.at[idx, 'Earnings'])
            hours = float(df.at[idx, 'Hours'])
            cost_code = float(df.at[idx, 'Cost Code'])
            employee = df.at[idx, 'Employee Name']
            earn_type = df.at[idx, 'Earn Type']

            if hours > 0:
                actual_rate = earnings / hours
            else:
                continue

            if cost_code in high_codes:
                threshold = high_codes[cost_code]['threshold']
                name = high_codes[cost_code]['name']
                low_code = high_codes[cost_code]['low_code']

                if actual_rate < threshold:
                    misclassifications.append({
                        'employee': employee,
                        'earn_type': earn_type,
                        'current_code': int(cost_code),
                        'should_be_code': int(low_code) if low_code else 'UNKNOWN',
                        'actual_rate': actual_rate,
                        'threshold': threshold,
                        'earnings_total': earnings,
                        'trade': name
                    })

            elif cost_code in low_codes:
                threshold = low_codes[cost_code]['threshold']
                name = low_codes[cost_code]['name']
                high_code = low_codes[cost_code]['high_code']

                if actual_rate >= threshold:
                    misclassifications.append({
                        'employee': employee,
                        'earn_type': earn_type,
                        'current_code': int(cost_code),
                        'should_be_code': int(high_code),
                        'actual_rate': actual_rate,
                        'threshold': threshold,
                        'earnings_total': earnings,
                        'trade': name
                    })

        except (ValueError, KeyError, TypeError):
            continue

    return df, misclassifications


def reclassify_drive_time(df):
    """
    Reclassify drive time (DRIVE, DROVT) to lower wage class codes.
    """
    class_code_map = get_class_code_mapping()
    drive_mask = df['Earn Type'].str.upper().str.strip().isin(['DRIVE', 'DROVT'])
    reclassifications = []

    for idx in df[drive_mask].index:
        original_code = df.at[idx, 'Cost Code']

        try:
            original_code_num = float(original_code)
        except:
            continue

        if original_code_num in class_code_map:
            new_code = class_code_map[original_code_num]
            df.at[idx, 'Cost Code'] = new_code

            reclassifications.append({
                'employee': df.at[idx, 'Employee Name'],
                'original_code': int(original_code_num),
                'new_code': new_code,
                'earn_type': df.at[idx, 'Earn Type'],
                'earnings': df.at[idx, 'Earnings']
            })

    return df, reclassifications


def apply_employee_specific_corrections(df):
    """
    Apply specific employee class code corrections.
    """
    corrections = []

    employee_rules = {
        'Kidwell , Austin': {
            'correct_code': 8810,
            'reason': 'Office/clerical duties only',
            'description': 'Clerical Office Employees (NOC)'
        }
    }

    for employee_name, rule in employee_rules.items():
        employee_mask = df['Employee Name'].str.strip() == employee_name.strip()

        if employee_mask.any():
            original_codes = df.loc[employee_mask, 'Cost Code'].unique()

            for original_code in original_codes:
                try:
                    if float(original_code) != rule['correct_code']:
                        code_mask = employee_mask & (df['Cost Code'] == original_code)
                        count = code_mask.sum()
                        total_earnings = df.loc[code_mask, 'Earnings'].sum()

                        df.loc[code_mask, 'Cost Code'] = rule['correct_code']

                        corrections.append({
                            'employee': employee_name,
                            'original_code': int(float(original_code)),
                            'corrected_code': rule['correct_code'],
                            'reason': rule['reason'],
                            'count': count,
                            'earnings': total_earnings
                        })
                except (ValueError, TypeError):
                    continue

    return df, corrections


def classify_wage_type(earn_type):
    """
    Classify earn type into Excel columns.
    """
    if pd.isna(earn_type):
        return "OTHER"

    earn_type = str(earn_type).upper().strip()

    regular_types = ['REG', 'VAC', 'BON', 'SUPP', 'SICK', 'DBA', 'DRIVE', 'OSAL', 'SAL', 'PWREG']
    overtime_types = ['OVT', 'DROVT', 'PWOT']
    doubletime_types = ['DBL']

    if earn_type in regular_types:
        return "REGULAR"
    elif earn_type in overtime_types:
        return "OVERTIME"
    elif earn_type in doubletime_types:
        return "DOUBLETIME"
    else:
        return "OTHER"


def read_csv_with_encoding(file_path):
    """
    Read CSV file with automatic encoding detection.
    """
    encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1', 'utf-16']

    for encoding in encodings:
        try:
            df = pd.read_csv(file_path, encoding=encoding)
            return df
        except (UnicodeDecodeError, UnicodeError):
            continue

    try:
        df = pd.read_csv(file_path, encoding='utf-8', errors='ignore')
        return df
    except Exception as e:
        raise ValueError(f"Unable to read CSV file with any encoding: {e}")


def process_csv_data(csv_path):
    """
    Process CSV data to group by employee and summarize wages by type.
    """
    print(f"Reading and processing CSV data: {os.path.basename(csv_path)}")
    df = read_csv_with_encoding(csv_path)
    print(f"Loaded {len(df)} rows of raw data")

    # Filter out summary rows
    df_filtered = df[~df['Job No'].isna() & (df['Job No'] != "")].copy()
    df_filtered = df_filtered[~df_filtered['Job Description'].str.contains('TOTAL', case=False, na=False)]

    print(f"After filtering: {len(df_filtered)} detail rows")

    # Apply corrections
    df_filtered, employee_corrections = apply_employee_specific_corrections(df_filtered)
    df_filtered, wage_misclassifications = validate_class_code_by_wage(df_filtered)
    df_filtered, reclassifications = reclassify_drive_time(df_filtered)

    # Parse employee names
    df_filtered[['First Name', 'Last Name']] = df_filtered['Employee Name'].apply(
        lambda x: pd.Series(parse_employee_name(x))
    )

    # Classify wage types
    df_filtered['Wage Type'] = df_filtered['Earn Type'].apply(classify_wage_type)

    # Group by employee and class code
    grouped = df_filtered.groupby(['Employee Number', 'Employee Name', 'First Name', 'Last Name', 'Cost Code'], dropna=False).agg({
        'Earnings': 'sum',
        'Exposure': 'sum',
    }).reset_index()

    # Pivot to get wage types as separate columns
    wage_summary = df_filtered.groupby(['Employee Number', 'Employee Name', 'First Name', 'Last Name', 'Cost Code', 'Wage Type'], dropna=False).agg({
        'Earnings': 'sum'
    }).reset_index()

    wage_pivot = wage_summary.pivot_table(
        index=['Employee Number', 'Employee Name', 'First Name', 'Last Name', 'Cost Code'],
        columns='Wage Type',
        values='Earnings',
        aggfunc='sum',
        fill_value=0
    ).reset_index()

    # Merge with main grouped data
    result = grouped.merge(
        wage_pivot,
        on=['Employee Number', 'Employee Name', 'First Name', 'Last Name', 'Cost Code'],
        how='left'
    )

    # Fill NaN values
    for col in ['REGULAR', 'OVERTIME', 'DOUBLETIME', 'OTHER']:
        if col not in result.columns:
            result[col] = 0
        else:
            result[col] = result[col].fillna(0)

    print(f"\nProcessed into {len(result)} employee/class code combinations")

    # Calculate total earnings from SOURCE DATA
    total_source_earnings = df_filtered['Earnings'].sum()
    print(f"\nTotal Earnings from source CSV: ${total_source_earnings:,.2f}")

    return result, total_source_earnings


def import_formatted_data_to_excel(processed_df, excel_path, output_dir, total_source_earnings, start_row=23, pay_period=None):
    """
    Import formatted data into Excel template.
    """
    print(f"\nLoading Excel template: {os.path.basename(excel_path)}")
    wb = load_workbook(excel_path)
    ws = wb["Payroll Entry"]

    # Unprotect the worksheet
    if ws.protection.sheet:
        ws.protection.sheet = False
        print("Worksheet protection removed")

    # Set column widths
    ws.column_dimensions['E'].width = 12

    # Write G12 - Gross Wages from Register
    ws.cell(row=12, column=7, value=total_source_earnings)
    print(f"G12 (Gross Wages from Register) set to: ${total_source_earnings:,.2f}")

    # Update pay period date if provided
    if pay_period:
        try:
            pay_period_date = datetime.strptime(pay_period, "%Y%m%d")
            pay_period_start = pay_period_date - timedelta(days=6)

            ws.cell(row=9, column=7, value=pay_period_date)
            ws.cell(row=10, column=7, value=pay_period_start)
            ws.cell(row=11, column=7, value=pay_period_date)

            print(f"Pay period dates updated:")
            print(f"  G9 (Report date): {pay_period_date.strftime('%Y-%m-%d')}")
            print(f"  G10 (Period start date): {pay_period_start.strftime('%Y-%m-%d')}")
            print(f"  G11 (Period end date): {pay_period_date.strftime('%Y-%m-%d')}")
        except ValueError:
            print(f"Warning: Invalid pay period format '{pay_period}'")

    # Clear old template data
    print(f"Clearing old data from rows {start_row} to 400...")
    for row in range(start_row, 401):
        for col in range(1, 17):
            cell = ws.cell(row=row, column=col)
            cell.value = None

    print(f"Starting data import at row {start_row}...")

    current_row = start_row

    for idx, row_data in processed_df.iterrows():
        # Column A: Employee Number
        employee_number = row_data.get('Employee Number', '')
        try:
            employee_number = int(float(str(employee_number).strip())) if pd.notna(employee_number) and employee_number != '' else ''
        except (ValueError, TypeError):
            employee_number = str(employee_number).strip() if pd.notna(employee_number) else ''

        ws.cell(row=current_row, column=1, value=employee_number)
        ws.cell(row=current_row, column=2, value=row_data['First Name'])
        ws.cell(row=current_row, column=3, value=row_data['Last Name'])
        ws.cell(row=current_row, column=4, value="CA")

        # Column E: Class Code - Ensure all digits display properly
        class_code = row_data['Cost Code']
        if pd.notna(class_code):
            # Convert to int to ensure it's numeric, preserve all digits
            class_code_value = int(float(class_code))
            class_code_cell = ws.cell(row=current_row, column=5, value=class_code_value)
            class_code_cell.number_format = '0'  # Use integer format to match CLI version

        # Wage columns
        ws.cell(row=current_row, column=6, value=round(row_data['REGULAR'], 2))
        ws.cell(row=current_row, column=7, value=round(row_data['OVERTIME'], 2))
        ws.cell(row=current_row, column=8, value=round(row_data['DOUBLETIME'], 2))

        current_row += 1

        if idx % 25 == 0 and idx > 0:
            print(f"  Processed {idx} records...")

    print(f"  Processed all {len(processed_df)} records")

    # Calculate totals
    total_regular = processed_df['REGULAR'].sum()
    total_overtime = processed_df['OVERTIME'].sum()
    total_doubletime = processed_df['DOUBLETIME'].sum()
    grand_total = total_regular + total_overtime + total_doubletime

    print(f"\nSUMMARY TOTALS")
    print(f"Regular Wages: ${total_regular:,.2f}")
    print(f"Overtime (1.5x): ${total_overtime:,.2f}")
    print(f"Double Time (2x): ${total_doubletime:,.2f}")
    print(f"GRAND TOTAL EARNINGS: ${grand_total:,.2f}")

    # Save the workbook
    if pay_period:
        time_part = datetime.now().strftime("%H%M%S")
        output_filename = f"Workers_Comp_{pay_period}_{time_part}.xlsx"
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"Workers_Comp_{timestamp}.xlsx"

    output_path = os.path.join(output_dir, output_filename)

    print(f"\nSaving formatted workbook to: {output_filename}")
    wb.save(output_path)
    wb.close()

    print(f"\n[SUCCESS] Data imported and formatted successfully!")
    print(f"Output file: {output_path}")

    return output_path, {
        'regular': total_regular,
        'overtime': total_overtime,
        'doubletime': total_doubletime,
        'grand_total': grand_total,
        'record_count': len(processed_df)
    }


def generate_standalone_armorpro_report(armorpro_csv, template_path, output_dir, pay_period):
    """
    Generate ArmorPro-only report using same template as combined report.

    Args:
        armorpro_csv: Path to the ArmorPro CSV file
        template_path: Path to Excel template file
        output_dir: Directory for output file
        pay_period: Pay period in YYYYMMDD format

    Returns:
        str: Path to generated ArmorPro Excel file
        dict: Totals dictionary with earnings breakdown
    """
    print(f"=== GENERATING ARMORPRO STANDALONE REPORT ===")
    print(f"Processing ArmorPro CSV: {os.path.basename(armorpro_csv)}")

    # Process the ArmorPro CSV data using existing function
    processed_data, total_source_earnings = process_csv_data(armorpro_csv)

    print(f"ArmorPro processed data: {len(processed_data)} records")
    print(f"ArmorPro total earnings: ${total_source_earnings:,.2f}")

    # Load Excel template
    print(f"Loading Excel template: {os.path.basename(template_path)}")
    wb = load_workbook(template_path)
    ws = wb["Payroll Entry"]

    # Unprotect the worksheet
    if ws.protection.sheet:
        ws.protection.sheet = False
        print("Worksheet protection removed")

    # Set column widths
    ws.column_dimensions['E'].width = 12

    # Write G12 - Gross Wages from Register
    ws.cell(row=12, column=7, value=total_source_earnings)
    print(f"G12 (Gross Wages from Register) set to: ${total_source_earnings:,.2f}")

    # Update pay period date if provided
    if pay_period:
        try:
            pay_period_date = datetime.strptime(pay_period, "%Y%m%d")
            pay_period_start = pay_period_date - timedelta(days=6)

            ws.cell(row=9, column=7, value=pay_period_date)
            ws.cell(row=10, column=7, value=pay_period_start)
            ws.cell(row=11, column=7, value=pay_period_date)

            print(f"Pay period dates updated:")
            print(f"  G9 (Report date): {pay_period_date.strftime('%Y-%m-%d')}")
            print(f"  G10 (Period start date): {pay_period_start.strftime('%Y-%m-%d')}")
            print(f"  G11 (Period end date): {pay_period_date.strftime('%Y-%m-%d')}")
        except ValueError:
            print(f"Warning: Invalid pay period format '{pay_period}'")

    # Clear old template data
    start_row = 23
    print(f"Clearing old data from rows {start_row} to 400...")
    for row in range(start_row, 401):
        for col in range(1, 17):
            cell = ws.cell(row=row, column=col)
            cell.value = None

    print(f"Starting ArmorPro data import at row {start_row}...")

    current_row = start_row

    for idx, row_data in processed_data.iterrows():
        # Column A: Employee Number
        employee_number = row_data.get('Employee Number', '')
        try:
            employee_number = int(float(str(employee_number).strip())) if pd.notna(employee_number) and employee_number != '' else ''
        except (ValueError, TypeError):
            employee_number = str(employee_number).strip() if pd.notna(employee_number) else ''

        ws.cell(row=current_row, column=1, value=employee_number)
        ws.cell(row=current_row, column=2, value=row_data['First Name'])
        ws.cell(row=current_row, column=3, value=row_data['Last Name'])
        ws.cell(row=current_row, column=4, value="CA")

        # Column E: Class Code - Ensure all digits display properly
        class_code = row_data['Cost Code']
        if pd.notna(class_code):
            # Convert to int to ensure it's numeric, preserve all digits
            class_code_value = int(float(class_code))
            class_code_cell = ws.cell(row=current_row, column=5, value=class_code_value)
            class_code_cell.number_format = '0'  # Use integer format to match CLI version

        # Wage columns
        ws.cell(row=current_row, column=6, value=round(row_data['REGULAR'], 2))
        ws.cell(row=current_row, column=7, value=round(row_data['OVERTIME'], 2))
        ws.cell(row=current_row, column=8, value=round(row_data['DOUBLETIME'], 2))

        current_row += 1

        if idx % 25 == 0 and idx > 0:
            print(f"  Processed {idx} ArmorPro records...")

    print(f"  Processed all {len(processed_data)} ArmorPro records")

    # Calculate totals
    total_regular = processed_data['REGULAR'].sum()
    total_overtime = processed_data['OVERTIME'].sum()
    total_doubletime = processed_data['DOUBLETIME'].sum()
    grand_total = total_regular + total_overtime + total_doubletime

    print(f"ARMORPRO SUMMARY TOTALS")
    print(f"Regular Wages: ${total_regular:,.2f}")
    print(f"Overtime (1.5x): ${total_overtime:,.2f}")
    print(f"Double Time (2x): ${total_doubletime:,.2f}")
    print(f"GRAND TOTAL EARNINGS: ${grand_total:,.2f}")

    # Save with standard filename format (subtle suffix to avoid conflicts with combined reports)
    if pay_period:
        time_part = datetime.now().strftime("%H%M%S")
        output_filename = f"Workers_Comp_{pay_period}_{time_part}_AP.xlsx"
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"Workers_Comp_{timestamp}_AP.xlsx"

    output_path = os.path.join(output_dir, output_filename)

    print(f"Saving formatted workbook to: {output_filename}")
    wb.save(output_path)
    wb.close()

    print(f"[SUCCESS] ArmorPro data imported and formatted successfully!")
    print(f"ArmorPro output file: {output_path}")

    return output_path, {
        'regular': total_regular,
        'overtime': total_overtime,
        'doubletime': total_doubletime,
        'grand_total': grand_total,
        'record_count': len(processed_data)
    }
